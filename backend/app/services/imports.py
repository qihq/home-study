import importlib.util
import mimetypes
import zipfile
from dataclasses import dataclass
from pathlib import Path

from app.services.words import normalize_word


class ParserUnavailable(Exception):
    code = 'PARSER_UNAVAILABLE'


@dataclass(frozen=True)
class ParsedImport:
    items: list[dict[str, str]]
    warnings: list[str]


def _line_items(text: str) -> list[dict[str, str]]:
    seen: set[str] = set()
    items = []
    for number, line in enumerate(text.splitlines(), start=1):
        display = ' '.join(line.strip().split())
        normalized = normalize_word(display)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        items.append({'display_text': display, 'normalized_text': normalized, 'source_location': f'line:{number}'})
    return items


def parse_import_file(path: Path, content_type: str | None = None) -> ParsedImport:
    suffix = path.suffix.casefold()
    if suffix in {'.txt', '.csv'} or content_type == 'text/plain':
        return ParsedImport(_line_items(path.read_text(encoding='utf-8-sig')), [])
    required_module = {'.xlsx': 'openpyxl', '.docx': 'docx', '.pdf': 'pypdf'}.get(suffix)
    if required_module and importlib.util.find_spec(required_module) is None:
        raise ParserUnavailable(f'{required_module} is not installed')
    if suffix == '.xlsx':
        from openpyxl import load_workbook
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except zipfile.BadZipFile as error:
            raise ParserUnavailable('openpyxl could not read the XLSX file') from error
        lines = []
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for column_index, value in enumerate(row, start=1):
                    if isinstance(value, str): lines.append((f'{sheet.title}!R{row_index}C{column_index}', value))
        return ParsedImport(_items_with_locations(lines), [])
    if suffix == '.docx':
        from docx import Document
        document = Document(path)
        return ParsedImport(_items_with_locations([(f'paragraph:{index}', paragraph.text) for index, paragraph in enumerate(document.paragraphs, start=1)]), [])
    if suffix == '.pdf':
        from pypdf import PdfReader
        reader = PdfReader(path)
        return ParsedImport(_items_with_locations([(f'page:{index}', page.extract_text() or '') for index, page in enumerate(reader.pages, start=1)]), ['扫描 PDF 需要安装 OCR 后才能识别。'])
    raise ValueError('UNSUPPORTED_IMPORT_TYPE')


def _items_with_locations(lines: list[tuple[str, str]]) -> list[dict[str, str]]:
    seen: set[str] = set(); result = []
    for location, raw in lines:
        for line in raw.splitlines():
            display = ' '.join(line.strip().split()); normalized = normalize_word(display)
            if normalized and normalized not in seen:
                seen.add(normalized); result.append({'display_text': display, 'normalized_text': normalized, 'source_location': location})
    return result
